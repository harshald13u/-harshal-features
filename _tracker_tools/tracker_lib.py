#!/usr/bin/env python3
"""
tracker_lib.py — the ONE source of truth logic for the Harshal Dasani media tracker.

Design rule (the permanent fix):
  * The "All" sheet is the SINGLE source of truth.
  * Per-publication sheets and the 3 HTML snapshots are DERIVED — never hand-edited.
  * Every add/repair goes through this library, so no chat can re-introduce drift.

Public API:
  load(path) / save(wb, path)
  load_taxonomy()
  canonical_url(u)         -> normalized URL (storage + dedup)
  article_id(u)            -> stable id for dedup (or None)
  clean_heading(h)         -> heading with slug-junk removed / apostrophes restored
  validate(wb, tax)        -> list[(severity, check, detail)]   (severity: ERROR|WARN)
  normalize(wb, tax)       -> dict summary of idempotent repairs applied (mutates wb)
  add_rows(wb, tax, rows)  -> dict {added, skipped}             (rows: list of dicts)
  snapshot_rows(wb)        -> list[dict] ready to embed as EMBEDDED_SNAPSHOT_ROWS
"""
import re, json, os, html, collections
from datetime import datetime, date

HEADER = ["Date","Heading Of The Article","Language","Publication","Link","Topic","Journalist"]
NONPUB = {"All","Special Features","Journalist Contacts"}
HERE = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------- taxonomy
def load_taxonomy(path=None):
    with open(path or os.path.join(HERE,"taxonomy.json"), encoding="utf-8") as f:
        return json.load(f)

# ---------------------------------------------------------------- urls / ids
_TRACK = re.compile(r'^(utm_|igsh|fbclid|gclid|ref_|spm)', re.I)
def canonical_url(u):
    u = (u or "").strip()
    if not u: return u
    u = u.split("#")[0]
    # drop tracking query params, keep meaningful ones
    if "?" in u:
        base, q = u.split("?", 1)
        keep = [p for p in q.split("&") if p and not _TRACK.match(p.split("=")[0])]
        u = base + ("?" + "&".join(keep) if keep else "")
    # host normalisations
    u = re.sub(r'^https?://m\.economictimes\.com', 'https://economictimes.indiatimes.com', u)
    u = re.sub(r'^https?://m\.livemint\.com', 'https://www.livemint.com', u)
    u = re.sub(r'/amp-(\d+)\.html', r'-\1.html', u)          # livemint amp
    u = re.sub(r'/amp/1?$', '', u)                            # ndtvprofit /amp or /amp/1
    u = re.sub(r'/amp/?$', '', u)                             # trailing /amp/
    return u

def article_id(u):
    u = (u or "").lower()
    for pat in (r'articleshow/(\d+)', r'/article-(\d+)', r'-(\d{6,})\.(?:html|cms|ece)',
                r'-(\d{7,})(?:/|$)', r'/reel/([a-z0-9_-]{6,})', r'/reel/(\d{6,})',
                r'/(cm[a-z0-9]{18,})(?:/|$)'):
        m = re.search(pat, u)
        if m: return m.group(1)
    return None

# ---------------------------------------------------------------- headings
_JUNK_SUFFIX = [
    (re.compile(r'\s+Articleshow\s+[a-z0-9]{5,}\s*$', re.I), ''),
    (re.compile(r'\s+\d{5,}\s+20\d2\s+\d{2}\s+\d{2}\s*$'), ''),   # trailing id + y m d
    (re.compile(r'\s+\d{6,}(\s+\d)?\s*$'), ''),                    # trailing id (+single digit)
    (re.compile(r'\.pdf\b', re.I), ''),
]
_APOS = [(re.compile(r'\bWhat S\b'),"What's"),(re.compile(r'\bHere S\b'),"Here's"),
         (re.compile(r'\bDay S\b'),"Day's"),(re.compile(r'\b([A-Z][a-z]+) S\b'),r"\1's"),
         (re.compile(r'\bIndias\b'),"India's"),(re.compile(r'\bVedantas\b'),"Vedanta's"),
         (re.compile(r'\bNvidias\b'),"Nvidia's"),(re.compile(r'\bCountrys\b'),"Country's"),
         (re.compile(r'\bModis\b'),"Modi's"),(re.compile(r'\bTrumps\b'),"Trump's")]
def clean_heading(h):
    h = html.unescape(h or "")
    h = re.sub(r'<[^>]+>', '', h)              # strip tags
    for rx, rep in _JUNK_SUFFIX: h = rx.sub(rep, h)
    for rx, rep in _APOS: h = rx.sub(rep, h)
    h = re.sub(r'\s{2,}', ' ', h).strip()
    return h

def heading_issues(h):
    out=[]
    if not (h or "").strip(): out.append("blank")
    if re.search(r'\bArticleshow\s+[a-z0-9]{5,}\b', h or "", re.I): out.append("articleshow-junk")
    if re.search(r'\.pdf\b', h or "", re.I): out.append("pdf-in-heading")
    if re.search(r'\s\d{6,}(\s\d)?\s*$', h or ""): out.append("trailing-id")
    if re.search(r'&(amp|quot|#x?\d+|nbsp|rsquo|lsquo|ldquo|rdquo|#39);', h or ""): out.append("html-entity")
    if re.search(r'<[^>]+>', h or ""): out.append("html-tag")
    if h and (h!=h.strip() or '  ' in h): out.append("whitespace")
    if re.search(r'\b[A-Z][a-z]+ S\b|\bWhat S\b|\bHere S\b', h or ""): out.append("apostrophe-stripped")
    return out

# ---------------------------------------------------------------- io
def load(path):
    import openpyxl
    return openpyxl.load_workbook(path)
def save(wb, path):
    wb.save(path)

def _dkey(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return date(1900,1,1)

def _rows(ws):
    return [list(r) for r in ws.iter_rows(min_row=2, values_only=True) if any(x is not None and str(x).strip() for x in r)]

# ---------------------------------------------------------------- validate
def validate(wb, tax):
    issues=[]
    E=lambda c,d: issues.append(("ERROR",c,d)); W=lambda c,d: issues.append(("WARN",c,d))
    if 'All' not in wb.sheetnames: return [("ERROR","structure","no All sheet")]
    ws=wb['All']
    hdr=[c.value for c in ws[1]]
    if hdr[:7]!=HEADER: issues.append(("ERROR","header",f"got {hdr[:7]}"))
    rows=_rows(ws)
    today=date.today()
    langs=set(tax["languages"]); topics=set(tax["topics"]); aliases=tax["topic_aliases"]
    seen_url=collections.Counter(); seen_id={}
    for i,r in enumerate(rows, start=2):
        d,h,lang,pub,link,topic,jour = (list(r)+[None]*7)[:7]
        # date
        if not isinstance(d,(datetime,date)): issues.append(("ERROR","date",f"r{i} not a date: {d!r}"))
        else:
            dd=d.date() if isinstance(d,datetime) else d
            if dd>today: issues.append(("ERROR","date",f"r{i} future {dd}"))
            if dd<date(2023,1,1): issues.append(("WARN","date",f"r{i} pre-2023 {dd}"))
        # heading
        for iss in heading_issues(h or ""): issues.append(("ERROR","heading",f"r{i} {iss}: {(h or '')[:50]}"))
        # link
        lk=(link or "").strip()
        if not lk: issues.append(("ERROR","link",f"r{i} blank"))
        elif not re.match(r'https?://',lk): issues.append(("ERROR","link",f"r{i} not http: {lk[:40]}"))
        else:
            seen_url[lk]+=1
            aid=article_id(lk)
            if aid:
                if aid in seen_id and seen_id[aid]!=lk: issues.append(("ERROR","dup-id",f"r{i} id {aid} also {seen_id[aid][:45]}"))
                seen_id.setdefault(aid,lk)
        # vocab
        if (lang or "").strip() not in langs: issues.append(("ERROR","language",f"r{i} {lang!r}"))
        if not (pub or "").strip(): issues.append(("ERROR","publication",f"r{i} blank"))
        t=(topic or "").strip()
        if not t: issues.append(("ERROR","topic",f"r{i} blank"))
        elif t not in topics: issues.append(("ERROR","topic",f"r{i} '{t}' not in taxonomy (alias->{aliases.get(t,'?')})"))
        if not (jour or "").strip(): issues.append(("ERROR","journalist",f"r{i} blank"))
    for u,c in seen_url.items():
        if c>1: issues.append(("ERROR","dup-url",f"x{c} {u[:60]}"))
    # sort order
    ooo=sum(1 for i in range(len(rows)-1) if _dkey(rows[i][0])<_dkey(rows[i+1][0]))
    if ooo: issues.append(("ERROR","sort",f"{ooo} out-of-order pairs in All"))
    # date format
    bad_fmt=sum(1 for r in range(2,ws.max_row+1) if ws.cell(r,1).number_format!='DD-MMM-YYYY')
    if bad_fmt: issues.append(("WARN","date-format",f"{bad_fmt} cells not DD-MMM-YYYY"))
    # per-pub consistency
    by_pub=collections.defaultdict(list)
    for r in rows: by_pub[(r[3] or '').strip()].append((r[4] or '').strip())
    data_sheets=[s for s in wb.sheetnames if s not in NONPUB]
    for pub,links in by_pub.items():
        if not pub: continue
        sh=next((s for s in wb.sheetnames if s.strip()==pub), None)
        if sh is None: issues.append(("ERROR","sheet-missing",f"no sheet for {pub!r}")); continue
        slinks=[(c[4].value or '').strip() if c[4].value else '' for c in wb[sh].iter_rows(min_row=2)]
        slinks=[x for x in slinks if x]
        if sorted(slinks)!=sorted(links): issues.append(("ERROR","sheet-mismatch",f"{pub!r}: All={len(links)} sheet={len(slinks)}"))
    for s in data_sheets:
        if s.strip() not in by_pub: issues.append(("ERROR","orphan-sheet",f"{s!r} has no rows in All"))
    return issues

# ---------------------------------------------------------------- normalize (idempotent repair)
def normalize(wb, tax):
    from openpyxl.utils import get_column_letter
    ch=collections.Counter()
    ws=wb['All']
    # header
    for i,v in enumerate(HEADER, start=1):
        if ws.cell(1,i).value!=v: ws.cell(1,i,value=v); ch['header']+=1
    rows=_rows(ws)
    aliases=tax["topic_aliases"]; pub_aliases=tax["publication_aliases"]
    out=[]
    for r in rows:
        r=(list(r)+[None]*7)[:7]
        d,h,lang,pub,link,topic,jour=r
        # coerce string date
        if isinstance(d,str):
            for fmt in ("%Y-%m-%d","%d-%b-%Y","%d-%b-%y","%d/%m/%Y"):
                try: d=datetime.strptime(d.strip(),fmt); ch['date-coerced']+=1; break
                except: pass
        nh=clean_heading(h or "")
        if nh!=(h or ""): ch['heading']+=1
        nlang=(lang or "").strip()
        npub=(pub or "").strip(); npub=pub_aliases.get(npub,npub)
        if npub!=(pub or "").strip(): ch['publication-alias']+=1
        nt=(topic or "").strip(); nt=aliases.get(nt,nt)
        if nt!=(topic or "").strip(): ch['topic-alias']+=1
        njour=(jour or "").strip()
        out.append([d,nh,nlang,npub,(link or "").strip(),nt,njour])
    # dedup exact URLs (keep first / newest since sorted later) and same-id
    seen=set(); seen_id=set(); dedup=[]
    out.sort(key=lambda r:_dkey(r[0]), reverse=True)   # newest first; keep first occurrence
    for r in out:
        u=r[4]; aid=article_id(u)
        if u in seen: ch['dup-url-removed']+=1; continue
        if aid and aid in seen_id: ch['dup-id-removed']+=1; continue
        seen.add(u);
        if aid: seen_id.add(aid)
        dedup.append(r)
    dedup.sort(key=lambda r:_dkey(r[0]), reverse=True)
    # rewrite All
    if ws.max_row>1: ws.delete_rows(2, ws.max_row-1)
    for r in dedup: ws.append(r)
    for rr in range(2, ws.max_row+1): ws.cell(rr,1).number_format='DD-MMM-YYYY'
    ws.auto_filter.ref=f"A1:{get_column_letter(7)}{ws.max_row}"
    # rebuild per-pub sheets from All
    by_pub=collections.defaultdict(list)
    for r in dedup: by_pub[(r[3] or '').strip()].append(r)
    for pub,prows in by_pub.items():
        if not pub: continue
        sh=next((wb[s] for s in wb.sheetnames if s.strip()==pub), None)
        if sh is None: sh=wb.create_sheet(pub); ch['sheet-created']+=1
        for i,v in enumerate(HEADER, start=1): sh.cell(1,i,value=v)
        if sh.max_row>1: sh.delete_rows(2, sh.max_row-1)
        for r in prows: sh.append(r)
        for rr in range(2, sh.max_row+1): sh.cell(rr,1).number_format='DD-MMM-YYYY'
        sh.auto_filter.ref=f"A1:{get_column_letter(7)}{sh.max_row}"
    # drop orphan sheets (no rows in All, and not protected)
    for s in [s for s in wb.sheetnames if s not in NONPUB and s.strip() not in by_pub]:
        del wb[s]; ch['orphan-sheet-removed']+=1
    return dict(ch)

# ---------------------------------------------------------------- add
def add_rows(wb, tax, rows):
    """rows: list of dicts with keys Date(datetime|YYYY-MM-DD), Heading, Language, Publication, Link, Topic, Journalist"""
    from openpyxl.utils import get_column_letter
    ws=wb['All']
    existing={canonical_url((r[4] or '')) for r in ws.iter_rows(min_row=2,values_only=True)}
    existing_id={article_id((r[4] or '')) for r in ws.iter_rows(min_row=2,values_only=True)} - {None}
    aliases=tax["topic_aliases"]; pub_aliases=tax["publication_aliases"]
    topics=set(tax["topics"]); langs=set(tax["languages"])
    added=0; skipped=[]
    def ins(sh,row):
        sh.insert_rows(2)
        for c,v in enumerate(row,1):
            cell=sh.cell(2,c,value=v)
            if c==1: cell.number_format='DD-MMM-YYYY'
        sh.auto_filter.ref=f"A1:{get_column_letter(7)}{sh.max_row}"
    for d in rows:
        link=canonical_url(d.get("Link",""))
        aid=article_id(link)
        if link in existing or (aid and aid in existing_id):
            skipped.append(("dup",link)); continue
        topic=(d.get("Topic") or "").strip(); topic=aliases.get(topic,topic)
        if topic not in topics: skipped.append(("bad-topic:"+topic,link)); continue
        lang=(d.get("Language") or "English").strip()
        if lang not in langs: skipped.append(("bad-language:"+lang,link)); continue
        pub=(d.get("Publication") or "").strip(); pub=pub_aliases.get(pub,pub)
        dt=d.get("Date")
        if isinstance(dt,str):
            dt=datetime.strptime(dt.strip(),"%Y-%m-%d")
        row=[dt, clean_heading(d.get("Heading","")), lang, pub, link, topic, (d.get("Journalist") or "Aggregator Desk").strip()]
        ins(ws,row)
        sh=next((wb[s] for s in wb.sheetnames if s.strip()==pub), None)
        if sh is None:
            sh=wb.create_sheet(pub)
            for i,v in enumerate(HEADER,1): sh.cell(1,i,value=v)
        ins(sh,row)
        existing.add(link);
        if aid: existing_id.add(aid)
        added+=1
    # keep All sorted after adds
    normalize(wb, tax)
    return {"added":added,"skipped":skipped}

# ---------------------------------------------------------------- snapshot
def snapshot_rows(wb):
    ws=wb['All']
    def ds(v): return v.strftime('%Y-%m-%d') if isinstance(v,(datetime,date)) else (str(v).strip() if v else '')
    out=[]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not (r[0] or (len(r)>4 and r[4])): continue
        out.append({"Date":ds(r[0]),"Heading":r[1] or '',"Language":r[2] or '',"Publication":r[3] or '',
                    "Link":r[4] or '',"Topic":(r[5] if len(r)>5 and r[5] else ''),
                    "Journalist":(r[6] if len(r)>6 and r[6] else '')})
    return out
